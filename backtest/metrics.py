import os
import json
import math
import pandas as pd
import numpy as np
from datetime import datetime
from glob import glob
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from typing import Tuple, Dict

class PerformanceMetrics:
    """績效計算類別"""
    
    def __init__(self, config: dict, file_name: str = None):
        """
        初始化績效計算類別
        
        Args:
            config: 配置參數，包含：
                - risk_free_rate: 無風險利率
                - initial_balance: 初始資金
            file_name: 指定的交易記錄檔案名稱，如果為None則讀取backtest_log目錄下所有檔案
        """
        self.config = config
        self.risk_free_rate = config.get('risk_free_rate', 0.025)
        self.initial_balance = config.get('initial_balance', 10000)
        self.file_name = file_name
        
        # 策略名稱對照表
        self.strategy_name_map = {
            'trend_long': '順勢做多',
            'trend_short': '順勢做空',
            'mean_rev_long': '逆勢做多',
            'mean_rev_short': '逆勢做空',
            'MANUAL': '手動買賣'
        }
        
        # 盤勢名稱對照表
        self.market_condition_map = {
            'long': '多頭',
            'short': '空頭',
            'sideway': '盤整'
        }
        
        # 初始化數據目錄
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.log_dir = os.path.join(self.base_dir, "backtest_log")
        
    def _ensure_numeric_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        確保指定的數值欄位為正確的數值型別
        
        Args:
            df: 要處理的數據框
            
        Returns:
            pd.DataFrame: 處理後的數據框
        """
        # 定義需要轉換為數值的欄位
        numeric_columns = ['pnl', 'pnl_percentage', 'open_price', 'close_price', 
                          'open_amt', 'close_amt', 'open_size', 'close_size', 'margin']
        
        # 對存在的欄位進行數值轉換
        for col in numeric_columns:
            if col in df.columns:
                # 記錄轉換前的無效數據數量
                original_invalid_count = pd.to_numeric(df[col], errors='coerce').isna().sum()
                
                # 執行轉換
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
                # 記錄轉換後的無效數據數量
                final_invalid_count = df[col].isna().sum()
                
                # 如果發現新的無效值，表示有數據轉換失敗
                if final_invalid_count > original_invalid_count:
                    print(f"警告: 在欄位 '{col}' 中發現 {final_invalid_count - original_invalid_count} 個無法解析的數值。它們已被設置為 NaN。")
                
                # 將 NaN 值填充為 0，以避免計算錯誤，這是一個重要的修正
                df[col] = df[col].fillna(0)
                
        return df
        
    def load_trade_logs(self) -> pd.DataFrame:
        """
        載入交易記錄
        
        Returns:
            pd.DataFrame: 交易記錄數據框
        """
        all_records = []
        
        if self.file_name:
            # 讀取指定的檔案
            file_path = os.path.join(self.log_dir, self.file_name)
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"找不到指定的交易記錄檔案: {file_path}")
                
            with open(file_path, "r") as f:
                for line in f:
                    record = json.loads(line)
                    all_records.append(record)
        else:
            # 讀取目錄下所有檔案
            for file_path in glob(os.path.join(self.log_dir, "*.jsonl")):
                with open(file_path, "r") as f:
                    for line in f:
                        record = json.loads(line)
                        all_records.append(record)
        
        df = pd.DataFrame(all_records)
        
        # 確保數值欄位為正確的型別
        df = self._ensure_numeric_columns(df)
        
        if not df.empty:
            # 統一新增 date 欄位
            df['date'] = pd.to_datetime(df['open_time']).dt.date
            
            # 從 market_condition 提取 trend_filter (應為 list)
            df['trend_filter'] = df['market_condition'].apply(
                lambda x: x.get('trend_filter') if isinstance(x, dict) and x.get('trend_filter') else []
            )
            
            # 基於 trend_filter list 創建 tuple 以便分組 (保留順序和重複)
            df['trend_combination'] = df['trend_filter'].apply(tuple)
            
        return df
        
    def _round_metrics(self, data):
        """
        將結果中的浮點數四捨五入到小數點後兩位
        """
        if isinstance(data, dict):
            result = {}
            for key, value in data.items():
                if isinstance(value, (float, np.floating)):
                    result[key] = round(value, 2)
                else:
                    result[key] = value
            return result
        elif isinstance(data, pd.DataFrame):
            data_copy = data.copy()
            for col in data_copy.select_dtypes(include=['float']).columns:
                data_copy[col] = data_copy[col].round(2)
            return data_copy
        return data
        
    def calculate_daily_metrics(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
        """
        計算每日績效指標
        
        Args:
            df: 交易記錄數據框
            
        Returns:
            Tuple[pd.DataFrame, Dict]: 每日績效數據框和總體摘要
        """
        # 確保數值欄位為正確的型別
        df = self._ensure_numeric_columns(df)
        
        # 計算每日績效
        agg_metrics = df.groupby('date').agg({
            'pnl': ['sum', 'count', 'mean']
        }).reset_index()
        agg_metrics.columns = ['日期', '當日損益 (USDT)', '交易次數 (筆)', '平均損益 (USDT)']
        
        # 直接以最終順序建立DataFrame
        daily_metrics = agg_metrics[['日期', '交易次數 (筆)', '當日損益 (USDT)', '平均損益 (USDT)']].copy()
        
        # 計算累計損益
        daily_metrics['累計損益 (USDT)'] = daily_metrics['當日損益 (USDT)'].cumsum()
        
        # 計算總體績效指標
        total_trades = len(df)
        total_pnl = df['pnl'].sum()
        wins = df[df['pnl'] > 0]
        losses = df[df['pnl'] < 0]
        win_rate = len(wins) / len(df) if len(df) > 0 else 0
        
        # 計算總交易天數
        trading_days = (df['date'].max() - df['date'].min()).days + 1 if not df.empty else 0
        
        # 計算總回報率
        total_return = total_pnl / self.initial_balance
        
        # 計算年化報酬率 (複利, 365個交易日)
        annual_return = (1 + total_return) ** (365 / trading_days) - 1 if trading_days > 0 and (1 + total_return) > 0 else 0
        
        # 計算夏普比率
        daily_pnl = daily_metrics.set_index('日期')['當日損益 (USDT)']
        daily_returns = daily_pnl / self.initial_balance
        if len(daily_returns) > 1:
            annualized_volatility = np.std(daily_returns) * math.sqrt(365)
            sharpe_ratio = (annual_return - self.risk_free_rate) / annualized_volatility if annualized_volatility != 0 else 0
        else:
            sharpe_ratio = 0
            
        # 建立每日權益曲線並計算最大回撤
        daily_equity = self.initial_balance + daily_pnl.cumsum()
        # 確保初始資金被視為第一個峰值
        equity_curve = pd.concat([pd.Series([self.initial_balance]), daily_equity], ignore_index=True)
        peak_equity = equity_curve.expanding(min_periods=1).max()
        
        # 最大回撤 (USDT)
        max_drawdown_usdt = (peak_equity - equity_curve).max()

        # 最大回撤百分比 (相對於歷史高點資金)
        drawdown_pct_series = (equity_curve - peak_equity) / peak_equity
        max_drawdown_pct = abs(drawdown_pct_series.min()) * 100 if not drawdown_pct_series.empty else 0
        
        # 計算卡瑪比率
        calmar_ratio = annual_return / (max_drawdown_pct / 100) if max_drawdown_pct > 0 else 0
        
        # 計算其他指標
        avg_win = wins['pnl'].mean() if not wins.empty else 0
        avg_loss = abs(losses['pnl'].mean()) if not losses.empty else 0
        profit_loss_ratio = avg_win / avg_loss if avg_loss != 0 else 0
        total_profit = wins['pnl'].sum()
        total_loss = abs(losses['pnl'].sum())
        profit_factor = total_profit / total_loss if total_loss != 0 else 0
        
        # 計算平均持倉時間（分鐘）
        df['holding_time'] = (pd.to_datetime(df['close_time']) - pd.to_datetime(df['open_time'])).dt.total_seconds() / 60
        avg_holding_time = df['holding_time'].mean()
        
        # 建立總體摘要
        summary = {
            '總交易次數 (筆)': total_trades,
            '總盈虧 (USDT)': total_pnl,
            '勝率 (%)': win_rate * 100,
            '年化報酬率 (%)': annual_return * 100,
            '平均獲利 (USDT)': avg_win,
            '平均虧損 (USDT)': avg_loss,
            '盈虧比': profit_loss_ratio,
            '最大回撤 (USDT)': max_drawdown_usdt,
            '最大回撤 (%)': max_drawdown_pct,
            '獲利因子': profit_factor,
            '夏普比率': sharpe_ratio,
            '卡瑪比率': calmar_ratio,
            '平均持倉時間 (分鐘)': avg_holding_time
        }
        
        # 四捨五入結果
        daily_metrics = self._round_metrics(daily_metrics)
        summary = self._round_metrics(summary)
        
        return daily_metrics, summary
        
    def calculate_common_metrics(self, group: pd.DataFrame) -> dict:
        """
        計算共同的績效指標
        
        Args:
            group: 交易記錄數據框
            
        Returns:
            dict: 包含所有共同績效指標的字典
        """
        if group.empty:
            return {}
        
        # --- 核心邏輯重構：基於每日數據進行計算 ---
        
        # 1. 匯總每日Pnl - 'date' 欄位由 load_trade_logs 保證存在
        daily_pnl = group.groupby('date')['pnl'].sum()

        # 2. 計算基礎指標
        total_pnl = daily_pnl.sum()
        wins = group[group['pnl'] > 0]
        losses = group[group['pnl'] < 0]
        win_rate = len(wins) / len(group) if len(group) > 0 else 0
        
        # 計算總交易天數 (使用有交易的日子)
        if daily_pnl.empty:
            trading_days = 0
        else:
            trading_days = (daily_pnl.index.max() - daily_pnl.index.min()).days + 1
            
        # 3. 計算年化報酬率 (複利, 365個交易日)
        total_return = total_pnl / self.initial_balance
        annual_return = (1 + total_return) ** (365 / trading_days) - 1 if trading_days > 0 and (1 + total_return) > 0 else 0
        
        # 4. 計算夏普比率 (基於每日報酬率)
        daily_returns = daily_pnl / self.initial_balance
        if len(daily_returns) > 1:
            # 使用 ddof=0 計算樣本標準差，以匹配 np.std 的預設行為
            annualized_volatility = daily_returns.std(ddof=0) * math.sqrt(365)
            sharpe_ratio = (annual_return - self.risk_free_rate) / annualized_volatility if annualized_volatility != 0 else 0
        else:
            sharpe_ratio = 0
            
        # 5. 建立每日權益曲線並計算最大回撤
        daily_equity = self.initial_balance + daily_pnl.cumsum()
        # 確保初始資金被視為第一個峰值
        equity_curve = pd.concat([pd.Series([self.initial_balance]), daily_equity], ignore_index=True)
        peak_equity = equity_curve.expanding(min_periods=1).max()
        
        # 最大回撤 (USDT)
        max_drawdown_usdt = (peak_equity - equity_curve).max()

        # 最大回撤百分比 (相對於歷史高點資金)
        drawdown_pct_series = (equity_curve - peak_equity) / peak_equity
        max_drawdown_pct = abs(drawdown_pct_series.min()) * 100 if not drawdown_pct_series.empty else 0
        
        # 6. 計算卡瑪比率
        calmar_ratio = annual_return / (max_drawdown_pct / 100) if max_drawdown_pct > 0 else 0
        
        # --- 其他指標計算 ---
        
        total_trades = len(group)
        avg_win = wins['pnl'].mean() if not wins.empty else 0
        avg_loss = abs(losses['pnl'].mean()) if not losses.empty else 0
        profit_loss_ratio = avg_win / avg_loss if avg_loss != 0 else 0
        
        total_profit = wins['pnl'].sum()
        total_loss = abs(losses['pnl'].sum())
        profit_factor = total_profit / total_loss if total_loss != 0 else 0

        # 計算平均持倉時間
        group_copy = group.copy()
        group_copy['holding_time'] = (pd.to_datetime(group_copy['close_time']) - pd.to_datetime(group_copy['open_time'])).dt.total_seconds() / 60
        avg_holding_time = group_copy['holding_time'].mean()

        # 建立結果字典
        metrics = {
            "交易次數 (筆)": total_trades,
            "總盈虧 (USDT)": total_pnl,
            "勝率 (%)": win_rate * 100,
            "年化報酬率 (%)": annual_return * 100,
            "平均獲利 (USDT)": avg_win,
            "平均虧損 (USDT)": avg_loss,
            "盈虧比": profit_loss_ratio,
            "最大回撤 (USDT)": max_drawdown_usdt,
            "最大回撤 (%)": max_drawdown_pct,
            "獲利因子": profit_factor,
            "夏普比率": sharpe_ratio,
            "卡瑪比率": calmar_ratio,
            "平均持倉時間 (分鐘)": avg_holding_time
        }
        
        return self._round_metrics(metrics)
        
    def calculate_symbol_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        計算各交易對的績效指標
        
        Args:
            df: 交易記錄數據框
            
        Returns:
            pd.DataFrame: 交易對績效數據框
        """
        symbol_stats = []
        
        for symbol, group in df.groupby('symbol'):
            if group.empty:
                continue
                
            # 計算共同績效指標
            common_metrics = self.calculate_common_metrics(group)
            
            # 添加交易對特有的欄位
            symbol_stats.append({
                '交易對': symbol,
                **common_metrics
            })
            
        return pd.DataFrame(symbol_stats)
        
    def calculate_strategy_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        計算每個策略的績效指標
        
        Args:
            df: 交易記錄數據框
            
        Returns:
            pd.DataFrame: 策略績效數據框
        """
        if df.empty:
            return pd.DataFrame()
            
        grouped = df.groupby('strategy')
        
        results = []
        for strategy, group in grouped:
            common_metrics = self.calculate_common_metrics(group)
            
            # 使用策略名稱對照表
            strategy_name = self.strategy_name_map.get(strategy, strategy)
            
            result_row = {
                '策略': strategy_name,
                **common_metrics
            }
            results.append(result_row)
            
        strategy_df = pd.DataFrame(results)
        return self._round_metrics(strategy_df)

    def calculate_market_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        計算不同盤勢下的績效指標，支持27種組合
        
        Args:
            df: 交易記錄數據框
            
        Returns:
            pd.DataFrame: 盤勢績效數據框
        """
        if df.empty or 'trend_combination' not in df.columns:
            return pd.DataFrame()
            
        # 按盤勢組合和策略分組計算
        # 'trend_combination' 欄位已在 load_trade_logs 中創建 (為元組格式)
        grouped = df.groupby(['trend_combination', 'strategy'])
        
        results = []
        for (trend_combination, strategy), group in grouped:
            if not trend_combination:  # 跳過沒有盤勢標籤的交易
                continue

            common_metrics = self.calculate_common_metrics(group)
            
            # 直接使用元組的字串表示法作為盤勢組合名稱
            trend_str = str(trend_combination)
            strategy_name = self.strategy_name_map.get(strategy, strategy)

            result_row = {
                '盤勢組合': trend_str,
                '策略': strategy_name,
                **common_metrics
            }
            results.append(result_row)
            
        market_df = pd.DataFrame(results)
        return self._round_metrics(market_df)

    def export_to_excel(self, daily_df: pd.DataFrame, summary_dict: dict, 
                       symbol_df: pd.DataFrame, strategy_df: pd.DataFrame, 
                       market_df: pd.DataFrame, output_file: str = "回測績效分析報告.xlsx") -> None:
        """
        匯出績效報告到Excel
        
        Args:
            daily_df: 每日績效數據框
            summary_dict: 總體摘要字典
            symbol_df: 交易對績效數據框
            strategy_df: 策略績效數據框
            market_df: 盤勢績效數據框
            output_file: 輸出文件名
        """
        # 定義共同績效指標的順序
        common_columns = [
            '交易次數 (筆)', '總盈虧 (USDT)', '勝率 (%)', '年化報酬率 (%)', '平均獲利 (USDT)', '平均虧損 (USDT)', '盈虧比', 
            '最大回撤 (USDT)', '最大回撤 (%)', '獲利因子', '夏普比率', '卡瑪比率', '平均持倉時間 (分鐘)'
        ]
        
        # 重新排列交易對績效表格的欄位順序
        if not symbol_df.empty:
            symbol_specific_cols = ['交易對']
            symbol_df = symbol_df[symbol_specific_cols + [col for col in common_columns if col in symbol_df.columns]]
        
        # 重新排列策略績效表格的欄位順序
        if not strategy_df.empty:
            strategy_specific_cols = ['策略']
            strategy_df = strategy_df[strategy_specific_cols + [col for col in common_columns if col in strategy_df.columns]]
        
        # 重新排列盤勢績效表格的欄位順序
        if not market_df.empty:
            market_specific_cols = ['盤勢組合', '策略']
            market_df = market_df[market_specific_cols + [col for col in common_columns if col in market_df.columns]]
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 總體摘要
                summary_df = pd.DataFrame([summary_dict]).T.reset_index().rename(columns={"index": "指標", 0: "數值"})
                summary_df.to_excel(writer, sheet_name="總體摘要", index=False)
                summary_sheet = writer.sheets["總體摘要"]
                # 設置列寬
                summary_sheet.column_dimensions['A'].width = 30  
                summary_sheet.column_dimensions['B'].width = 15  
                # 設置對齊方式
                for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 每日績效
                daily_df.to_excel(writer, sheet_name="每日績效", index=False)
                daily_sheet = writer.sheets["每日績效"]
                # 設置列寬
                daily_sheet.column_dimensions['A'].width = 15  
                daily_sheet.column_dimensions['B'].width = 25  
                daily_sheet.column_dimensions['C'].width = 25  
                daily_sheet.column_dimensions['D'].width = 25  
                daily_sheet.column_dimensions['E'].width = 25  
                # 設置對齊方式
                for row in daily_sheet.iter_rows(min_row=1, max_row=daily_sheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 交易對績效
                symbol_df.to_excel(writer, sheet_name="交易對績效", index=False)
                symbol_sheet = writer.sheets["交易對績效"]
                # 設置列寬
                for idx, col in enumerate(symbol_df.columns):
                    col_letter = get_column_letter(idx + 1)
                    symbol_sheet.column_dimensions[col_letter].width = 20
                symbol_sheet.column_dimensions['N'].width = 25
                # 設置對齊方式
                for row in symbol_sheet.iter_rows(min_row=1, max_row=symbol_sheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 策略分析
                strategy_df.to_excel(writer, sheet_name="策略分析", index=False)
                strategy_sheet = writer.sheets["策略分析"]
                # 設置列寬
                for idx, col in enumerate(strategy_df.columns):
                    col_letter = get_column_letter(idx + 1)
                    strategy_sheet.column_dimensions[col_letter].width = 20
                strategy_sheet.column_dimensions['N'].width = 25
                # 設置對齊方式
                for row in strategy_sheet.iter_rows(min_row=1, max_row=strategy_sheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 盤勢分析
                market_df.to_excel(writer, sheet_name="盤勢分析", index=False)
                market_sheet = writer.sheets["盤勢分析"]
                # 設置列寬
                for idx, col in enumerate(market_df.columns):
                    col_letter = get_column_letter(idx + 1)
                    market_sheet.column_dimensions[col_letter].width = 20
                market_sheet.column_dimensions['A'].width = 30
                market_sheet.column_dimensions['O'].width = 25
                # 設置對齊方式
                for row in market_sheet.iter_rows(min_row=1, max_row=market_sheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 設置標題行格式
                for sheet in [daily_sheet, summary_sheet, symbol_sheet, strategy_sheet, market_sheet]:
                    for cell in sheet[1]:
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        cell.font = Font(bold=True)

        except PermissionError:
            print(f"無法寫入文件 {output_file}，請確保文件未被其他程序使用。")
            # 嘗試使用不同的文件名
            new_output_file = f"回測績效分析報告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            print(f"嘗試使用新的文件名: {new_output_file}")
            self.export_to_excel(daily_df, summary_dict, symbol_df, strategy_df, market_df, new_output_file)
            
    def run(self) -> None:
        """執行績效分析"""
        # 載入交易記錄
        df = self.load_trade_logs()
        if df.empty:
            print("找不到任何交易資料。")
            return
            
        # 計算各項績效指標
        daily_df, summary = self.calculate_daily_metrics(df)
        symbol_df = self.calculate_symbol_metrics(df)
        strategy_df = self.calculate_strategy_metrics(df)
        market_df = self.calculate_market_metrics(df)
        
        # 匯出績效報告到根目錄
        root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        output_file = os.path.join(root_dir, f"回測績效分析報告_{self.file_name}.xlsx")
        self.export_to_excel(daily_df, summary, symbol_df, strategy_df, market_df, output_file)
        print(f"分析報告已匯出至: {output_file}")