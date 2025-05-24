import os
import json
import math
import pandas as pd
from datetime import datetime
from glob import glob
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(__file__)
DATA_FOLDER = os.path.join(BASE_DIR, "trade_log")
RISK_FREE_RATE = 0.025
LEVERAGE = 5

# 策略名稱對照表
STRATEGY_NAME_MAP = {
    'trend_long': '順勢做多',
    'trend_short': '順勢做空',
    'mean_rev_long': '逆勢做多',
    'mean_rev_short': '逆勢做空',
    'MANUAL': '手動買賣'
}

def load_jsonl_files(folder):
    all_records = []
    for file_path in glob(os.path.join(folder, "*.jsonl")):
        with open(file_path, "r") as f:
            for line in f:
                record = json.loads(line)
                record['file_date'] = os.path.basename(file_path).replace(".jsonl", "")
                all_records.append(record)
    return pd.DataFrame(all_records)

def std(values):
    if len(values) <= 1:
        return 0
    mean = sum(values) / len(values)
    squared_diffs = [(x - mean) ** 2 for x in values]
    return math.sqrt(sum(squared_diffs) / (len(values) - 1))

def compute_metrics(df):
    # 確保數據類型正確
    df['pnl_percentage'] = pd.to_numeric(df['pnl_percentage'], errors='coerce')
    df['pnl'] = pd.to_numeric(df['pnl'], errors='coerce')
    df['open_price'] = pd.to_numeric(df['open_price'], errors='coerce')
    df['open_amt'] = pd.to_numeric(df['open_amt'], errors='coerce')
    
    # 使用文件日期作為主要日期
    df['date'] = pd.to_datetime(df['file_date']).dt.date
    
    # 計算每筆交易的實際投入資金（考慮槓桿）
    df['investment'] = (df['open_amt'] * df['open_price']) / LEVERAGE
    
    # 計算每日總投入資金和總盈虧
    daily_summary = df.groupby('date').agg(
        每日交易次數=('symbol', 'count'),
        每日盈虧=('pnl', 'sum'),
        總投入資金=('investment', 'sum')
    ).reset_index()
    
    # 計算每日報酬率
    daily_summary['每日報酬率'] = (daily_summary['每日盈虧'] / daily_summary['總投入資金']) * 100
    daily_summary['每日報酬率'] = round(daily_summary['每日報酬率'], 2)
    
    # 重命名欄位
    daily_summary = daily_summary.rename(columns={
        'date': '日期',
        '每日交易次數': '每日交易次數 (筆)',
        '每日盈虧': '每日盈虧 (USDT)',
        '每日報酬率': '每日報酬率 (%)'
    })
    
    # 計算總報酬率
    total_investment = daily_summary['總投入資金'].sum()
    total_pnl = daily_summary['每日盈虧 (USDT)'].sum()
    total_return = total_pnl / total_investment
    
    # 計算交易天數
    trading_days = (daily_summary['日期'].max() - daily_summary['日期'].min()).days + 1
    
    # 計算年化報酬率
    annualized_return = (1 + total_return) ** (365 / trading_days) - 1
    
    # 計算日報酬率的標準差
    daily_returns = daily_summary['每日報酬率 (%)'].values / 100
    std_daily_return = std(daily_returns)
    
    # 計算年化波動率
    annualized_volatility = std_daily_return * math.sqrt(365 / trading_days)
    
    # 計算夏普比率
    # 根據交易天數調整無風險利率
    daily_risk_free_rate = (1 + RISK_FREE_RATE) ** (1/365) - 1
    sharpe_ratio = (annualized_return - RISK_FREE_RATE) / annualized_volatility if annualized_volatility != 0 else 0
    
    # 計算最大回撤（使用臨時變數）
    temp_cumulative = total_investment + daily_summary['每日盈虧 (USDT)'].cumsum()
    temp_peak = temp_cumulative.cummax()
    temp_drawdown = (temp_peak - temp_cumulative) / temp_peak
    max_drawdown = temp_drawdown.max()
    
    # 計算勝率和其他指標
    wins = df[df['pnl'] > 0]
    losses = df[df['pnl'] < 0]
    win_rate = len(wins) / len(df) if len(df) > 0 else 0
    avg_win = wins['pnl'].mean() if not wins.empty else 0
    avg_loss = abs(losses['pnl'].mean()) if not losses.empty else 0
    
    # 計算盈虧比
    profit_loss_ratio = avg_win / avg_loss if avg_loss != 0 else 0
    
    # 計算期望報酬
    expected_return = (win_rate * avg_win) - ((1 - win_rate) * avg_loss)
    
    # 計算最大連續虧損次數
    df['is_loss'] = df['pnl'] < 0
    df['loss_streak'] = df['is_loss'].groupby((~df['is_loss']).cumsum()).cumsum()
    max_loss_streak = df['loss_streak'].max() if 'loss_streak' in df else 0
    
    # 計算平均持倉時間
    df['duration_minutes'] = (df['close_time'] - df['open_time']) / 1000 / 60
    avg_holding = df['duration_minutes'].mean()
    
    # 移除不需要的欄位
    daily_summary = daily_summary.drop(columns=['總投入資金'])
    
    summary = {
        '總交易次數 (筆)': len(df),
        '平均每日交易次數 (筆)': round(df.groupby('date').size().mean(), 2),
        '年化報酬率 (%)': round(annualized_return * 100, 2),
        '年化波動率 (%)': round(annualized_volatility * 100, 2),
        '夏普比率': round(sharpe_ratio, 2),
        '勝率 (%)': round(win_rate * 100, 2),
        '平均獲利 (USDT)': round(avg_win, 2),
        '平均虧損 (USDT)': round(avg_loss, 2),
        '盈虧比': round(profit_loss_ratio, 2),
        '期望報酬 (USDT)': round(expected_return, 2),
        '最大回撤 (%)': round(max_drawdown * 100, 2),
        '最大連續虧損次數 (筆)': max_loss_streak,
        '平均持倉時間 (分鐘)': round(avg_holding, 2),
        '交易期間 (天)': trading_days
    }
    
    return daily_summary, summary

def compute_strategy_breakdown(df):
    strategy_stats = []

    for strategy, group in df.groupby('strategy'):
        if group.empty:
            continue

        # 轉換策略名稱為中文
        strategy_name = STRATEGY_NAME_MAP.get(strategy, strategy)

        group['date'] = pd.to_datetime(group['file_date']).dt.date
        group['duration_minutes'] = (group['close_time'] - group['open_time']) / 1000 / 60
        wins = group[group['pnl'] > 0]
        losses = group[group['pnl'] < 0]

        # 計算每日總投入資金和總盈虧（使用暫時變數）
        daily_pnl = group.groupby('date')['pnl'].sum()
        daily_investment = group.groupby('date').apply(lambda x: sum(x['open_amt'] * x['open_price']) / LEVERAGE)
        
        # 計算總報酬率
        total_investment = daily_investment.sum()
        total_pnl = daily_pnl.sum()
        total_return = total_pnl / total_investment
        
        # 計算交易天數
        trading_days = (group['date'].max() - group['date'].min()).days + 1
        
        # 計算年化報酬率
        annualized_return = (1 + total_return) ** (365 / trading_days) - 1
        
        # 計算日報酬率的標準差
        daily_returns = (daily_pnl / daily_investment).values
        std_daily_return = std(daily_returns)
        
        # 計算年化波動率
        annualized_volatility = std_daily_return * math.sqrt(365 / trading_days)
        
        # 計算夏普比率
        sharpe_ratio = (annualized_return - RISK_FREE_RATE) / annualized_volatility if annualized_volatility != 0 else 0

        group['cum_pnl'] = group['pnl'].cumsum()
        max_drawdown = (group['cum_pnl'] - group['cum_pnl'].cummax()).min()

        # 計算策略的盈虧比和期望報酬
        win_rate = len(wins) / len(group) if len(group) > 0 else 0
        avg_win = wins['pnl'].mean() if not wins.empty else 0
        avg_loss = abs(losses['pnl'].mean()) if not losses.empty else 0
        profit_loss_ratio = avg_win / avg_loss if avg_loss != 0 else 0
        expected_return = (win_rate * avg_win) - ((1 - win_rate) * avg_loss)

        strategy_stats.append({
            '策略名稱': strategy_name,
            '總交易次數 (筆)': len(group),
            '勝率 (%)': round(win_rate * 100, 2),
            '平均獲利 (USDT)': round(avg_win, 2),
            '平均虧損 (USDT)': round(avg_loss, 2),
            '盈虧比': round(profit_loss_ratio, 2),
            '期望報酬 (USDT)': round(expected_return, 2),
            '年化報酬率 (%)': round(annualized_return * 100, 2),
            '年化波動率 (%)': round(annualized_volatility * 100, 2),
            '夏普比率': round(sharpe_ratio, 2),
            '最大回撤 (USDT)': round(max_drawdown, 2),
            '平均持倉時間 (分鐘)': round(group['duration_minutes'].mean(), 2)
        })

    return pd.DataFrame(strategy_stats)

def export_to_excel(daily_df, summary_dict, strategy_df, output_file="績效分析報告.xlsx"):
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 每日績效
            daily_df.to_excel(writer, sheet_name="每日績效", index=False)
            daily_sheet = writer.sheets["每日績效"]
            # 設置列寬
            daily_sheet.column_dimensions['A'].width = 15  # 日期
            daily_sheet.column_dimensions['B'].width = 25  # 每日交易次數
            daily_sheet.column_dimensions['C'].width = 25  # 每日盈虧
            daily_sheet.column_dimensions['D'].width = 25  # 每日報酬率
            # 設置對齊方式
            for row in daily_sheet.iter_rows(min_row=1, max_row=daily_sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 總體摘要
            summary_df = pd.DataFrame([summary_dict]).T.reset_index().rename(columns={"index": "指標", 0: "數值"})
            summary_df.to_excel(writer, sheet_name="總體摘要", index=False)
            summary_sheet = writer.sheets["總體摘要"]
            # 設置列寬
            summary_sheet.column_dimensions['A'].width = 30  # 指標
            summary_sheet.column_dimensions['B'].width = 15  # 數值
            # 設置對齊方式
            for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 策略分析
            strategy_df.to_excel(writer, sheet_name="策略分析", index=False)
            strategy_sheet = writer.sheets["策略分析"]
            # 設置列寬
            for idx, col in enumerate(strategy_df.columns):
                col_letter = get_column_letter(idx + 1)
                strategy_sheet.column_dimensions[col_letter].width = 20
            strategy_sheet.column_dimensions['L'].width = 30
            # 設置對齊方式
            for row in strategy_sheet.iter_rows(min_row=1, max_row=strategy_sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 設置標題行格式
            for sheet in [daily_sheet, summary_sheet, strategy_sheet]:
                for cell in sheet[1]:
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.font = cell.font.copy(bold=True)

    except PermissionError:
        print(f"無法寫入文件 {output_file}，請確保文件未被其他程序使用。")
        # 嘗試使用不同的文件名
        new_output_file = f"績效分析報告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"嘗試使用新的文件名: {new_output_file}")
        with pd.ExcelWriter(new_output_file, engine='openpyxl') as writer:
            # 重複上述格式設置
            daily_df.to_excel(writer, sheet_name="每日績效", index=False)
            daily_sheet = writer.sheets["每日績效"]
            daily_sheet.column_dimensions['A'].width = 15
            daily_sheet.column_dimensions['B'].width = 25
            daily_sheet.column_dimensions['C'].width = 25
            daily_sheet.column_dimensions['D'].width = 25
            for row in daily_sheet.iter_rows(min_row=1, max_row=daily_sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            summary_df = pd.DataFrame([summary_dict]).T.reset_index().rename(columns={"index": "指標", 0: "數值"})
            summary_df.to_excel(writer, sheet_name="總體摘要", index=False)
            summary_sheet = writer.sheets["總體摘要"]
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 15
            for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            strategy_df.to_excel(writer, sheet_name="策略分析", index=False)
            strategy_sheet = writer.sheets["策略分析"]
            for idx, col in enumerate(strategy_df.columns):
                col_letter = get_column_letter(idx + 1)
                strategy_sheet.column_dimensions[col_letter].width = 20
            strategy_sheet.column_dimensions['L'].width = 30
            for row in strategy_sheet.iter_rows(min_row=1, max_row=strategy_sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 設置標題行格式
            for sheet in [daily_sheet, summary_sheet, strategy_sheet]:
                for cell in sheet[1]:
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.font = cell.font.copy(bold=True)

if __name__ == "__main__":
    df = load_jsonl_files(DATA_FOLDER)
    if df.empty:
        print("找不到任何交易資料。")
    else:
        daily_df, summary = compute_metrics(df)
        strategy_df = compute_strategy_breakdown(df)
        export_to_excel(daily_df, summary, strategy_df)
        print("分析報告已匯出。")
